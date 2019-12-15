import time

import openpyxl
import win32com.client as win32
import xlwings as xw

fileName = r'H:\新建文件夹\3、建安-新 - 副本 - 副本.xlsx'
linkColumn = 22
totalSheetName = "汇总"
detailsheetName = "合并明细（带合同号）"


def excelProccessXlwings():
    app = xw.App(visible=True, add_book=False)
    # 连接到excel
    workbook = app.books.open(fileName)  # 连接excel文件
    # 保存
    workbook.save(".\\3、建安-新 - 副本 - 副本_new1.xlsx")
    workbook.close()
    app.quit()
def excelProcess():
    print("当前时间戳为:", time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    excelFile = openpyxl.load_workbook(fileName, read_only=False, data_only=False)
    totalTable = excelFile["汇总"]
    totalTableRowCnt = totalTable.max_row  # 汇总行数
    detailTable = excelFile["合并明细（带合同号）"]
    detailTableRowCnt = totalTable.max_row  # 合并明细（带合同号）行数
    for rowIndex in range(totalTableRowCnt):
        # totalTable.cell(row=rowIndex + 1, column=linkColumn+2).value = '= HYPERLINK("{}", "{}")'.format(".\\3.建安图片","LinkName")
        print(totalTable.cell(row=rowIndex + 1, column=linkColumn).value)
    print("当前时间戳为:", time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    excelFile.save(".\\3、建安-新 - 副本 - 副本_new.xlsx")


def excelProccessWin32():
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    fileName = r'H:\新建文件夹\3、建安-新 - 副本 - 副本.xlsx'
    wb = excel.Workbooks.Open(fileName)


excelProccessXlwings()
